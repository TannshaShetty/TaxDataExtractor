import streamlit as st
import fitz  # PyMuPDF
from PIL import Image
import google.generativeai as genai
import os
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# Load environment variables
load_dotenv()
GOOGLE_API_KEY = os.getenv('GOOGLE_API_KEY')
genai.configure(api_key='AIzaSyC15hBMiMRDoF42JRuiHrCfrmC2VM6IKF8')

# Model Configuration
MODEL_CONFIG = {
    "temperature": 0.2,
    "top_p": 1,
    "top_k": 32,
    "max_output_tokens": 4096,
}

# Initialize Gemini Model
model = genai.GenerativeModel(model_name="gemini-1.5-flash", generation_config=MODEL_CONFIG)

# Set the background gradient using CSS with a darker overlay
st.markdown(
    """
    <style>
    body {
        background-image: url('https://coolbackgrounds.io/images/backgrounds/index/compute-ea4c57a4.png');
        background-size: cover;
        background-position: center;
        background-attachment: fixed;
        color: #fff;  /* Change text color to white for contrast */
        font-family: 'Arial', sans-serif;
    }

    .stApp {
        background: rgba(0, 0, 0, 0.5);  /* Darken the background with a black overlay */
    }
    </style>
    """, unsafe_allow_html=True
)

# Convert the first page of a PDF to an image
def pdf_to_image(pdf_path):
    try:
        doc = fitz.open(pdf_path)
        page = doc[0]
        pix = page.get_pixmap()
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        return img
    except Exception as e:
        raise ValueError(f"Error converting PDF to image: {e}")


# Generate AI response from the image
def gemini_output(image, system_prompt, user_prompt):
    try:
        img_byte_arr = io.BytesIO()
        image.save(img_byte_arr, format="JPEG")
        img_byte_arr.seek(0)

        input_prompt = [system_prompt, {"mime_type": "image/jpeg", "data": img_byte_arr.getvalue()}, user_prompt]
        response = model.generate_content(input_prompt)
        return response.text.strip()
    except Exception as e:
        raise ValueError(f"Error generating output from Gemini: {e}")


# Parse AI output into a structured DataFrame
def parse_invoice_data(delimited_data):
    try:
        rows = [row.strip().split(",") for row in delimited_data.strip().split("\n")]

        # Fixed headers for invoice data
        headers = ["Total Amount", "Base Amount", "Tax Amount", "Recipient Name", "Sender Name", "Invoice Date", "Invoice Number"]

        if len(rows) > 0 and all(x.lower() in headers[0].lower() for x in rows[0]):
            df = pd.DataFrame(rows[1:], columns=rows[0])  # Use AI headers if provided
        else:
            df = pd.DataFrame(rows, columns=headers)  # Use fixed headers

        return df
    except Exception as e:
        raise ValueError(f"Error parsing AI output: {e}")


# Save DataFrame to Excel with auto-adjusted column widths
def save_excel_with_autofit(df, output_path):
    try:
        df.to_excel(output_path, index=False, engine="openpyxl")

        # Auto-adjust column widths
        wb = load_workbook(output_path)
        ws = wb.active
        for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells) + 2
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length

        wb.save(output_path)

    except Exception as e:
        raise ValueError(f"Error saving Excel: {e}")


# Streamlit App
def app():
    st.title("📄 Invoice Data Extractor")

    uploaded_file = st.file_uploader("📂 Upload a PDF Invoice", type="pdf")

    if uploaded_file is not None:
        pdf_path = os.path.join("temp", uploaded_file.name)
        os.makedirs("temp", exist_ok=True)
        with open(pdf_path, "wb") as f:
            f.write(uploaded_file.getbuffer())

        try:
            img = pdf_to_image(pdf_path)
            st.image(img, caption="📸 Extracted PDF Page", use_column_width=True)

            system_prompt = "You are an expert in extracting structured invoice details from images."
            user_prompt = """Extract the following invoice details in a structured CSV format:
            total_amount,base_amount,tax_amount,recipient_name,sender_name,invoice_date,invoice_number
            Reply only with the values, separated by commas."""

            delimited_data = gemini_output(img, system_prompt, user_prompt)

            if delimited_data:
                df = parse_invoice_data(delimited_data)

                # Display formatted table with auto-adjusted column width
                st.subheader("📊 Extracted Invoice Data")
                st.dataframe(df.style.set_table_styles([{"selector": "td", "props": [("text-align", "left")]}]),
                             use_container_width=True)

                # Save as Excel with proper formatting
                output_path = f"temp/{uploaded_file.name}_invoice.xlsx"
                save_excel_with_autofit(df, output_path)

                with open(output_path, "rb") as f:
                    st.download_button("📂 Download Excel File", data=f, file_name=f"{uploaded_file.name}_invoice.xlsx")

        except Exception as e:
            st.error(f"⚠ Error: {e}")


if __name__ == "__main__":
    app()
